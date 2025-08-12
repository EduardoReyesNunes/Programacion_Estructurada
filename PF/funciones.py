import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import datetime

def borrarPantalla():
  os.system("cls")

def esperarTecla():
  input("\n\t\t ... ⚠️  Presione enter para continuar ⚠️ ...")

def menu_usurios():
   print("\n \t☕🍰 .:: Delicias and cofee® ::. 🍰☕\n\n\t\t📝 1.-  Registro  \n\t\t📂 2.-  Login \n\t\t⭕ 3.- Salir ")
   opcion=input("\n\t👉 Elige una opción: ").upper().strip() 
   return opcion

def admin():
   print("\n\t1️⃣  Agregar producto \n\t2️⃣  Agregar vendedor \n\t3️⃣  Mostrar productos \n\t4️⃣  Mostar vendedor \n\t5️⃣  Modificar producto \n\t6️⃣  Modificar vendedor \n\t7️⃣  Eliminar producto \n\t8️⃣  Eliminar vendedor \n\t9️⃣  Mostrar ventas por vendedor \n\t🔟 SALIR ")
   opcion = input("\n\t👉 Elige una opción: ").upper().strip()
   return opcion 

def vendedor():
   borrarPantalla()
   print("\n\t📄 .::  Menú Vendedor ::. 📄\n\n\t1️⃣  Crear venta \n\t2️⃣  Mostrar ventas \n\t3️⃣  Modificar \n\t4️⃣  Eliminar \n\t5️⃣  Salir ")
   opcion = input("\n\t👉 Elige una opción: ").upper().strip()
   return opcion 

def exportar_a_excel(nombre_archivo, datos, columnas, titulo):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = titulo

        # Estilo para el encabezado
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center")

        # Escribir encabezados
        for col_num, columna in enumerate(columnas, 1):
            celda = ws.cell(row=1, column=col_num, value=columna)
            celda.font = bold_font
            celda.alignment = center_alignment

        # Escribir datos
        for row_num, fila in enumerate(datos, 2):
            for col_num, valor in enumerate(fila, 1):
                ws.cell(row=row_num, column=col_num, value=valor)

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        # Guardar archivo
        fecha_actual = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        nombre_final = f"{nombre_archivo}_{fecha_actual}.xlsx"
        wb.save(nombre_final)
        print(f"\n\t✅ ¡Archivo '{nombre_final}' generado con éxito! ✅")
        return True
    
    except Exception as e:
        print(f"\n\t❌ Error al exportar a Excel: {e} ❌")
        return False